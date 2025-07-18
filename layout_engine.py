import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from openpyxl import Workbook
import os

def generate_layout_and_summary(csv_path, output_folder, sheet_width=48.0, sheet_height=96.0):
    df = pd.read_csv(csv_path)
    df = df.dropna(subset=['Width', 'Height'])
    df['Width'] = df['Width'].astype(float)
    df['Height'] = df['Height'].astype(float)
    df['Thickness'] = df['Thickness'].astype(float)
    df['Quantity'] = df['Quantity'].fillna(1).astype(int)
    df_expanded = df.loc[df.index.repeat(df['Quantity'])].copy().reset_index(drop=True)

    grouped = df_expanded.groupby('Thickness')

    os.makedirs(output_folder, exist_ok=True)

    for thickness, group in grouped:
        for strategy in ['row', 'column']:
            parts = group.to_dict('records')
            sheet_count = 1
            kerf = 0.25
            while parts:
                prev_count = len(parts)
                fig, ax = plt.subplots(figsize=(8.27, 11.69))  # A4 portrait
                ax.set_xlim(0, sheet_width)
                ax.set_ylim(0, sheet_height)
                ax.invert_yaxis()
                ax.set_aspect('equal')
                ax.set_title(f"Thickness {thickness} - Sheet {sheet_count} ({strategy})")

                x_cursor = y_cursor = 0
                row_max_height = col_max_width = 0

                remaining_parts = []

                for part in parts:
                    pw, ph = part['Width'], part['Height']
                    fit = True

                    if strategy == 'row':
                        if x_cursor + pw > sheet_width:
                            x_cursor = 0
                            y_cursor += row_max_height + kerf
                            row_max_height = 0
                        if y_cursor + ph > sheet_height:
                            fit = False
                    else:
                        if y_cursor + ph > sheet_height:
                            y_cursor = 0
                            x_cursor += col_max_width + kerf
                            col_max_width = 0
                        if x_cursor + pw > sheet_width:
                            fit = False

                    if fit:
                        rect = patches.Rectangle((x_cursor, y_cursor), pw, ph, linewidth=1, edgecolor='black', facecolor='lightgrey')
                        ax.add_patch(rect)
                        label = f"{part['Part Name']}\n{pw:.2f}x{ph:.2f}"
                        ax.text(x_cursor + pw / 2, y_cursor + ph / 2, label, ha='center', va='center', fontsize=6)
                        if strategy == 'row':
                            x_cursor += pw + kerf
                            row_max_height = max(row_max_height, ph)
                        else:
                            y_cursor += ph + kerf
                            col_max_width = max(col_max_width, pw)
                    else:
                        remaining_parts.append(part)

                layout_filename = os.path.join(output_folder, f"layout_thickness_{thickness}_sheet_{sheet_count}_{strategy}.png")
                plt.savefig(layout_filename, dpi=300, bbox_inches='tight')
                plt.close()

                if len(remaining_parts) == prev_count:
                    print(f"Warning: could not place remaining parts for thickness {thickness} using strategy '{strategy}'.")
                    break

                parts = remaining_parts
                sheet_count += 1

    summary_wb = Workbook()
    for thickness, group in grouped:
        sheet = summary_wb.create_sheet(title=f"Thickness {thickness}")
        sheet.append(["Part Name", "Width", "Height", "Thickness", "Material", "Quantity"])
        for _, row in group.iterrows():
            sheet.append([
                row["Part Name"],
                row["Width"],
                row["Height"],
                row["Thickness"],
                row.get("Material", ""),
                row["Quantity"]
            ])
    if "Sheet" in summary_wb.sheetnames:
        del summary_wb["Sheet"]
    summary_wb.save(os.path.join(output_folder, "cutlist_summary.xlsx"))
